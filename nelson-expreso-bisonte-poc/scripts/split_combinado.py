"""
split_combinado.py — Separar un Excel combinado (INICIAL + SISTEMA en sheets)
en dos archivos individuales con nombres estandarizados.

Uso:
    python3 split_combinado.py /path/al/IA_CONTADO_DDMM.xlsx /path/salida/

Genera:
    Inicial{DD}_{MM}.xlsx  — sheet INICIAL
    Sistema{DD}{MM}.xlsx   — sheet SISTEMA

Ejemplo:
    python3 split_combinado.py /home/server/.hermes/document_cache/IA_CONTADO_1906.xlsx \
        /home/server/proyectos/excel-merger-poc/test_data/
"""

import sys
import os
import openpyxl
from pathlib import Path


def split_combinado(src_path: str, dest_dir: str) -> None:
    src = Path(src_path)
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)

    # Extraer fecha del nombre si tiene formato *DDMM*
    stem = src.stem  # ej: "IA_CONTADO_1906"
    fecha = ""
    for part in stem.split("_"):
        if part.isdigit() and len(part) == 4:
            fecha = part  # ej: "1906"
            break

    dd = fecha[:2] if fecha else "XX"
    mm = fecha[2:] if fecha else "XX"

    wb_src = openpyxl.load_workbook(src, data_only=True)

    mapping = {
        "INICIAL": f"Inicial{dd}_{mm}.xlsx",
        "SISTEMA": f"Sistema{fecha}.xlsx",
    }

    for sheet_name, out_name in mapping.items():
        if sheet_name not in wb_src.sheetnames:
            print(f"[WARN] Sheet '{sheet_name}' no encontrada en {src.name}")
            continue

        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = sheet_name
        ws_src = wb_src[sheet_name]

        for row in ws_src.iter_rows():
            for cell in row:
                ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

        out_path = dest / out_name
        wb_new.save(out_path)
        print(f"Guardado: {out_path}  ({ws_src.max_row} filas x {ws_src.max_column} cols)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    split_combinado(sys.argv[1], sys.argv[2])

#!/usr/bin/env python3
"""
Re-arma expreso_bisonte_combinado.xlsx desde los sources originales.
Ejecutar cuando /tmp/excel-merger/ esté vacío (se borra con cada reinicio del server).

Uso:
    python3 rearmar_combinado.py
    # Luego subir al backend:
    curl -X POST http://localhost:9000/api/v1/excel/upload \
         -F file=@/tmp/expreso_ref/expreso_bisonte_combinado.xlsx

FUENTE ACTUALIZADA (jun 2026): todas las 4 sheets viven en un solo archivo.
"""
import os
import openpyxl

OUT_DIR = '/tmp/expreso_ref'
OUT = f'{OUT_DIR}/expreso_bisonte_combinado.xlsx'

# Fuente única con las 4 sheets (actualizada jun 2026)
SRC_SINGLE = '/home/server/brainstorming/2026-06-03-expreso-bisonte-diccionario-datos-pablo/entregables-xlsx/CONTADO_PABLO_RUIZ_origen.xlsx'

SHEET_MAPPING = {
    # src_name → dst_name
    'CDO SISTEMA':          'CDO Sistema',
    'PTE DE FACT SISTEMA':  'PTE de Fact Sistema',
    'CDO TRABAJADA':        'CDO TRABAJADA',
    'PF TRABAJADA':         'PF TRABAJADA',
}

os.makedirs(OUT_DIR, exist_ok=True)

src_wb = openpyxl.load_workbook(SRC_SINGLE, data_only=True)
combined = openpyxl.Workbook()
combined.remove(combined.active)

for src_name, dst_name in SHEET_MAPPING.items():
    ws_src = src_wb[src_name]
    ws_dst = combined.create_sheet(title=dst_name)
    for row in ws_src.iter_rows(values_only=True):
        ws_dst.append(row)
    print(f"  OK {dst_name:25s} ({ws_src.max_row} filas)")

combined.save(OUT)
print(f"\nGuardado: {OUT} ({os.path.getsize(OUT)} bytes)")
print("\nAhora subir al backend:")
print(f"  curl -X POST http://localhost:9000/api/v1/excel/upload -F file=@{OUT}")

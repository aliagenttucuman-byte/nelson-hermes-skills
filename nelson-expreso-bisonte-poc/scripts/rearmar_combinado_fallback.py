#!/usr/bin/env python3
"""
Rearmar combinado de Expreso Bisonte cuando rearmar_combinado.py no existe.

PITFALL: NO renombrar ws.title in-place sobre el wb origen — el rename muta el wb
y la iteración posterior ve nombres ya cambiados, dejando el combinado con solo 2
sheets (faltan las SISTEMA). Patrón correcto: copiar a un Workbook nuevo en orden
explícito.

Verificación: load_workbook(dst).sheetnames debe ser exactamente
['CDO Sistema','PTE de Fact Sistema','CDO TRABAJADA','PF TRABAJADA'].
Upload exitoso reporta rows≈406 y >17 columnas. 17 columnas → faltan SISTEMA.
"""
import os
import sys
from openpyxl import load_workbook, Workbook

SRC = '/home/server/brainstorming/2026-06-03-expreso-bisonte-diccionario-datos-pablo/entregables-xlsx/CONTADO_PABLO_RUIZ_origen.xlsx'
DST = '/tmp/excel-merger/expreso_bisonte_combinado.xlsx'

# new_name (lo que espera el backend) -> old_name (lo que tiene el origen)
REV_MAP = {
    'CDO Sistema': 'CDO SISTEMA',
    'PTE de Fact Sistema': 'PTE DE FACT SISTEMA',
    'CDO TRABAJADA': 'CDO TRABAJADA',
    'PF TRABAJADA': 'PF TRABAJADA',
}
ORDER = ['CDO Sistema', 'PTE de Fact Sistema', 'CDO TRABAJADA', 'PF TRABAJADA']


def main():
    os.makedirs(os.path.dirname(DST), exist_ok=True)
    src_wb = load_workbook(SRC)
    missing = [REV_MAP[n] for n in ORDER if REV_MAP[n] not in src_wb.sheetnames]
    if missing:
        print(f"ERROR: faltan sheets en origen: {missing}", file=sys.stderr)
        print(f"sheets disponibles: {src_wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    out = Workbook()
    out.remove(out.active)
    for new_name in ORDER:
        s = src_wb[REV_MAP[new_name]]
        t = out.create_sheet(new_name)
        for row in s.iter_rows(values_only=True):
            t.append(row)
    out.save(DST)

    final = load_workbook(DST).sheetnames
    print(f"sheets finales: {final}")
    assert final == ORDER, f"orden incorrecto: {final}"
    print(f"OK -> {DST}")


if __name__ == '__main__':
    main()

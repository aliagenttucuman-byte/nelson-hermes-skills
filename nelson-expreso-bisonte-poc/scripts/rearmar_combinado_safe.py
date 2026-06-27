#!/usr/bin/env python3
"""
Rearma /tmp/excel-merger/expreso_bisonte_combinado.xlsx con las 4 sheets exactas
que espera el backend. Usar este fallback cuando el rearmar_combinado.py del repo
falle por FileNotFoundError o cuando aparezca con menos de 4 sheets en el upload.

PITFALL HISTÓRICO (jun 2026) — NO USAR ESTE PATRÓN:
    wb = load_workbook(src)
    wb['CDO SISTEMA'].title = 'CDO Sistema'         # rename in-place
    wb['PTE DE FACT SISTEMA'].title = 'PTE de Fact Sistema'
    # delete sheets no mapeadas...
    wb.save(dst)

Si alguno de los nombres "destino" colisiona con una sheet existente del workbook,
openpyxl descarta silenciosamente UNA. Resultado: terminás con 2 sheets en vez
de 4 sin ningún error. El upload al backend reporta filas pero el split-system
falla porque faltan sheets.

FIX correcto (Workbook nuevo + iter_rows): este script.
"""
from openpyxl import load_workbook, Workbook
import os, sys, requests

SRC = '/home/server/brainstorming/2026-06-03-expreso-bisonte-diccionario-datos-pablo/entregables-xlsx/CONTADO_PABLO_RUIZ_origen.xlsx'
DST = '/tmp/excel-merger/expreso_bisonte_combinado.xlsx'
BACKEND = 'http://localhost:9000/api/v1/excel/upload'

REV_MAP = {
    'CDO Sistema':         'CDO SISTEMA',
    'PTE de Fact Sistema': 'PTE DE FACT SISTEMA',
    'CDO TRABAJADA':       'CDO TRABAJADA',
    'PF TRABAJADA':        'PF TRABAJADA',
}
ORDER = ['CDO Sistema', 'PTE de Fact Sistema', 'CDO TRABAJADA', 'PF TRABAJADA']


def rearmar(src=SRC, dst=DST):
    if not os.path.exists(src):
        raise FileNotFoundError(f"Origen no existe: {src}")
    os.makedirs(os.path.dirname(dst), exist_ok=True)

    src_wb = load_workbook(src)
    out = Workbook()
    out.remove(out.active)
    for new_name in ORDER:
        old_name = REV_MAP[new_name]
        if old_name not in src_wb.sheetnames:
            raise KeyError(f"Falta sheet '{old_name}' en origen. Tiene: {src_wb.sheetnames}")
        s = src_wb[old_name]
        t = out.create_sheet(new_name)
        for row in s.iter_rows(values_only=True):
            t.append(row)
    out.save(dst)

    final = load_workbook(dst).sheetnames
    assert final == ORDER, f"sheets finales incorrectos: {final}"
    print(f"OK rearmado: {dst}")
    print(f"   sheets: {final}")
    return dst


def upload(dst=DST, backend=BACKEND):
    with open(dst, 'rb') as f:
        r = requests.post(backend,
                          files={'file': (os.path.basename(dst), f,
                                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                          timeout=30)
    r.raise_for_status()
    d = r.json()
    print(f"OK upload: file_id={d.get('file_id')} rows={d.get('rows')}")
    return d


if __name__ == '__main__':
    rearmar()
    if '--upload' in sys.argv:
        upload()

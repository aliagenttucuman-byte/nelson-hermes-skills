#!/usr/bin/env python3
"""
Diagnóstico de diferencias entre INICIAL, SISTEMA y FINAL del merger Cobranzas Contado.
Uso: python3 diagnostico_merger_contado.py INICIAL.xlsx SISTEMA.xlsx [FINAL.xlsx]
"""
import sys
import openpyxl
from collections import Counter

def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f"  Sheets: {wb.sheetnames}, activa: {ws.title}")
    header_row = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip().lower() == 'nro':
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        print(f"  ERROR: no encontró header 'nro' en {path}")
        return [], []
    headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        nro = str(row.get('nro') or '').strip()
        if not nro or nro[:2] not in ('A.', 'B.'):
            continue
        row['nro'] = nro
        rows.append(row)
    return headers, rows

def get_estado(row):
    for k in row:
        if k.strip().upper() == 'ESTADO':
            return str(row[k] or '').strip().upper()
    return ''

def get_referente(row):
    for k in row:
        if k.strip().upper() == 'REFERENTE':
            return str(row[k] or '').strip().upper()
    return ''

def print_section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print('='*60)

if len(sys.argv) < 3:
    print("Uso: python3 diagnostico_merger_contado.py INICIAL.xlsx SISTEMA.xlsx [FINAL.xlsx]")
    sys.exit(1)

ini_path = sys.argv[1]
sis_path = sys.argv[2]
fin_path = sys.argv[3] if len(sys.argv) > 3 else None

print_section("INICIAL")
headers_ini, rows_ini = load_rows(ini_path)
estados_ini = Counter(get_estado(r) for r in rows_ini)
print(f"Total filas: {len(rows_ini)}")
print(f"Estados: {dict(estados_ini.most_common())}")
print(f"Headers: {headers_ini[:8]}")
dupes_ini = {k:v for k,v in Counter(r['nro'] for r in rows_ini).items() if v > 1}
if dupes_ini: print(f"⚠️  DUPLICADOS: {dupes_ini}")

print_section("SISTEMA")
headers_sis, rows_sis = load_rows(sis_path)
estados_sis = Counter(get_estado(r) for r in rows_sis)
print(f"Total filas: {len(rows_sis)}")
print(f"Estados: {dict(estados_sis.most_common())}")
dupes_sis = {k:v for k,v in Counter(r['nro'] for r in rows_sis).items() if v > 1}
if dupes_sis: print(f"⚠️  DUPLICADOS: {dupes_sis}")

# Cruce
dict_ini = {r['nro']: r for r in rows_ini}
dict_sis = {r['nro']: r for r in rows_sis}
nros_ini = set(dict_ini)
nros_sis = set(dict_sis)

existentes = nros_ini & nros_sis
nuevos     = nros_sis - nros_ini
eliminados = nros_ini - nros_sis

print_section("CRUCE INICIAL ↔ SISTEMA")
print(f"Existentes (INI & SIS): {len(existentes)}")
print(f"Nuevos     (SIS - INI): {len(nuevos)}")
print(f"Eliminados (INI - SIS): {len(eliminados)}")

# ED breakdown
ed_sis_nros = {r['nro'] for r in rows_sis if get_estado(r) == 'ED'}
ed_existentes = ed_sis_nros & existentes
ed_nuevos     = ed_sis_nros - nros_ini
ed_eliminados = {r['nro'] for r in rows_ini if get_estado(r) == 'ED'} - nros_sis

print(f"\nED existentes (van al FINAL): {len(ed_existentes)}")
print(f"ED nuevos     (van al FINAL): {len(ed_nuevos)}")
print(f"ED TOTAL esperado en FINAL:  {len(ed_existentes) + len(ed_nuevos)}")
print(f"ED del INICIAL eliminados:   {len(ed_eliminados)}")
if ed_eliminados:
    print(f"  → {sorted(ed_eliminados)}")

# Cambios de estado
print_section("CAMBIOS DE ESTADO (INICIAL → SISTEMA)")
cambios = []
for nro in existentes:
    e_ini = get_estado(dict_ini[nro])
    e_sis = get_estado(dict_sis[nro])
    if e_ini != e_sis:
        cambios.append((nro, e_ini, e_sis))

print(f"Total cambios de estado: {len(cambios)}")
cambios_a_ed = [(n, ei, es) for n, ei, es in cambios if es == 'ED']
cambios_de_ed = [(n, ei, es) for n, ei, es in cambios if ei == 'ED']
print(f"  → Cambiaron A ED:    {len(cambios_a_ed)}")
print(f"  → Cambiaron DESDE ED: {len(cambios_de_ed)}")

if cambios_a_ed:
    print("\nGuías que pasaron a ED (estas deben aparecer como ED en el FINAL):")
    for nro, ei, es in cambios_a_ed:
        ref_ini = get_referente(dict_ini[nro])
        succobro = str(dict_sis[nro].get('succobro') or '').strip().upper()
        print(f"  {nro}  {ei}→{es}  ref_ini={ref_ini or '(vacío)'}  succobro={succobro}")

# FINAL de Edith (opcional)
if fin_path:
    print_section(f"FINAL de Edith ({fin_path})")
    headers_fin, rows_fin = load_rows(fin_path)
    estados_fin = Counter(get_estado(r) for r in rows_fin)
    print(f"Total filas: {len(rows_fin)}")
    print(f"Estados: {dict(estados_fin.most_common())}")

    ed_fin_nros = {r['nro'] for r in rows_fin if get_estado(r) == 'ED'}
    ed_fin_ref  = {r['nro']: get_referente(r) for r in rows_fin if get_estado(r) == 'ED'}

    print(f"\nED en FINAL Edith:   {len(ed_fin_nros)}")
    print(f"ED esperado en FINAL: {len(ed_existentes) + len(ed_nuevos)}")

    faltantes = ed_fin_nros - (ed_existentes | ed_nuevos)
    sobrantes  = (ed_existentes | ed_nuevos) - ed_fin_nros
    if faltantes:
        print(f"\n⚠️  ED en Edith pero NO en merger: {sorted(faltantes)}")
    if sobrantes:
        print(f"⚠️  ED en merger pero NO en Edith: {sorted(sobrantes)}")

    # Por REFERENTE
    print_section("ED por REFERENTE (Edith vs merger esperado)")
    refs_edith = Counter(ed_fin_ref.values())
    print("Edith:", dict(refs_edith.most_common()))
